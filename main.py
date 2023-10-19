from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

# acao = input('Qual codigo da açao voce quer processar?').upper()
acao = 'bidi4'

with open(f'./dados/{acao}.txt', 'r') as arquivo_cotacao:
    # Pega todas as linhas do arquivo
    linhas = arquivo_cotacao.readlines()
    
    # formata linha a linha da forma que queremos
    linhas = [linha.replace('\n', '').split(';') for linha in linhas]

    #cria um planilha em memoria
    workbook = Workbook()

    # pega a atual ativa e muda o nome
    aba_atual = workbook.active
    aba_atual.title = "Dados"

    # criaçao do cabeçalho
    aba_atual.append(['DATA', 'COTAÇAO', 'BANDA INFERIOR', 'BANDA SUPERIOR'])

    indice = 2
    for linha in linhas:
        # DATA
        ano_mes_dia = linha[0].split(' ')[0].split('-')
        data = date(
            int(ano_mes_dia[0]),
            int(ano_mes_dia[1]),
            int(ano_mes_dia[2])
        )

        #COTACAO
        cotacao = float(linha[1])


        #Preenchendo as colunas dinamicamente
        aba_atual[f'A{indice}'] = data
        aba_atual[f'B{indice}'] = cotacao

        #BANDA INFERIOR
        aba_atual[f'C{indice}'] = f'=AVERAGE(B{indice}:B{indice+19}) - 2*STDEV(B{indice}:B{indice+19})'
        #BANDA SUPERIOR
        aba_atual[f'D{indice}'] = f'=AVERAGE(B{indice}:B{indice+19}) + 2*STDEV(B{indice}:B{indice+19}'

        indice += 1


    # NOVA ABA
    aba_grafico = workbook.create_sheet('Grafico')
    workbook.active = aba_grafico

    #mesclagem de celulas
    aba_grafico.merge_cells("A1:T2")

    # CRIAÇAO DO CABECALHO
    cabecalho = aba_grafico['A1'] #com a mesclagem toda essa a1 a t2 se tornaram a1
    cabecalho.font = Font(b=True, sz=18, color='FFFFFF')
    cabecalho.fill = PatternFill('solid', fgColor='333333')
    cabecalho.alignment =  Alignment(vertical='center', horizontal='center')
    cabecalho.value = f'Cotaçoes {acao}'


    # CRIAÇAO DO GRAFICO
    grafico = LineChart()
    grafico.width = 33.87
    grafico.height = 14.82
    grafico.title = 'Historico de Cotaçoes'
    grafico.x_axis.title = 'Data da Cotaçao'
    grafico.y_axis.title = 'Valor da Cotaçao'


    # Pegar dados de outra aba
    cotacoes_referencia = Reference(aba_atual, min_col=2, min_row=2, max_col=4, max_row=indice)
    datas_referencia = Reference(aba_atual, min_col=1, min_row=2, max_col=1, max_row=indice)

    grafico.add_data(cotacoes_referencia)
    grafico.set_categories(datas_referencia)

    aba_grafico.add_chart(grafico, 'A3')

    workbook.save('./planilha/dados.xlsx')